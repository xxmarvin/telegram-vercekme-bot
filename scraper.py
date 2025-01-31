# scraper.py
import os
import time
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

# -------------------------------------------------------------
# Ortak renkli metin fonksiyonları (isteğe bağlı, log için kullanılıyor)
def kirmizi_metin(text): return f"\033[38;2;255;0;0m{text}\033[0m"
def yesil_metin(text):   return f"\033[38;2;0;190;0m{text}\033[0m"
def sari_metin(text):   return f"\033[38;2;255;255;0m{text}\033[0m"
# -------------------------------------------------------------

def scrape_links(link: str, desired_links: int, output_csv: str = "links.csv") -> None:
    """
    main.py'deki mantık: Belirtilen link sayısı kadar href toplayarak links.csv'ye yazar.
    Parametreler:
        link          : İlk sayfadaki link (içinde page=1 vb. yer alan).
        desired_links : Kaç adet link çekileceği.
        output_csv    : Çıktı CSV dosyası (varsayılan: links.csv).
    """
    # Eğer output_csv sıfırdan oluşturulacaksa eskiyi silmek isteyebilirsiniz
    if os.path.exists(output_csv):
        os.remove(output_csv)

    linnk = link.replace("page=1", "page={page}")
    options = webdriver.ChromeOptions()
    options.add_argument("--headless")
    options.add_argument("--log-level=3")

    driver = webdriver.Chrome(options=options)
    collected_links = 0
    page = 1

    while collected_links < desired_links:
        driver.get(linnk.format(page=page))
        wait = WebDriverWait(driver, 10)
        elements = wait.until(EC.presence_of_all_elements_located((By.CSS_SELECTOR, "jv-result-summary-title a")))

        for element in elements:
            href_link = element.get_attribute("href")
            if href_link:
                with open(output_csv, "a", encoding="utf-8") as f:
                    f.write(href_link + "\n")
                print(kirmizi_metin(href_link))
                collected_links += 1
                if collected_links >= desired_links:
                    break

        if collected_links < desired_links:
            try:
                driver.execute_script('document.querySelector("#shared-pagination-next > a > span").click()')
                time.sleep(2)
                page += 1
                print(sari_metin(f"Sayfa: {page}"))
            except:
                print(kirmizi_metin("Tüm Sayfalar Çekildi!"))
                break
        else:
            break

    driver.quit()
    print(yesil_metin("Tüm linkler Çekildi! Şimdi veri kazma başlatılıyor..."))


def run_data_mining(
    links_csv: str = "links.csv",
    excel_file: str = "list.xlsx"
) -> int:
    """
    veri_kaz.py'deki mantık: links.csv'deki linkleri dolaşarak verileri list.xlsx'e kaydeder.
    Parametreler:
        links_csv  : Kaynak CSV (varsayılan: links.csv)
        excel_file : Çıktı Excel dosyası (varsayılan: list.xlsx)
    Return:
        basarili_islem : Kaç adet işlemin başarıyla tamamlandığı.
    """

    start_time = time.time()
    basarili_islem = 0

    # Excel Hazırlığı
    try:
        wb = load_workbook(excel_file)
        ws = wb.active
    except FileNotFoundError:
        wb = Workbook()
        ws = wb.active
        header = ["Email", "Vacancy", "Employer Name", "Contact Name",
                  "Mobile", "Company Name", "Address", "Industry"]
        ws.append(header)
        fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
        for cell in ws[1]:
            cell.fill = fill
        # Sütun genişlikleri
        for col in ws.columns:
            max_length = 40
            column = col[0].column_letter
            ws.column_dimensions[column].width = max_length

    # Selenium Ayarları
    chrome_options = Options()
    chrome_options.add_experimental_option('excludeSwitches', ["enable-automation", 'enable-logging'])
    chrome_options.add_argument('--disable-logging')
    chrome_options.add_argument('--log-level=3')
    chrome_options.add_argument('--disable-infobars')
    chrome_options.add_argument('--disable-extensions')
    chrome_options.add_argument("--disable-dev-shm-usage")
    chrome_options.add_argument('--disable-gpu')
    chrome_options.add_argument('--headless')  # Headless

    driver = webdriver.Chrome(options=chrome_options)
    wait = WebDriverWait(driver, 6)

    # CSV okunuyor
    df = pd.read_csv(links_csv, header=None, names=['link'])

    for index, row in df.iterrows():
        url = row['link']
        driver.get(url)
        time.sleep(2)

        # Çerez onayı - hata vermezse tıkla
        try:
            accept_btn = driver.find_element(By.CSS_SELECTOR, 'a.wt-ecl-button.wt-ecl-button--primary.cck-actions-button')
            accept_btn.click()
        except:
            pass

        try:
            # Bilgileri yakalama
            email_address = ""
            phone_number = ""
            vacancy = ""
            job_location = ""
            employer_name = ""
            job_sector = ""
            company_name = ""
            contact_name = ""

            # E-posta
            try:
                email_elem = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, 'a[href^="mailto:"]')))
                email_address = email_elem.text.strip()
            except:
                pass

            # Telefon
            try:
                phone_elem = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, 'li#jv-details-telNumber-0-0')))
                phone_number_raw = phone_elem.text
                phone_number = phone_number_raw.replace('Tel.:', '').strip()
            except:
                pass

            # Vacancy
            try:
                vacancy_spans = wait.until(
                    EC.presence_of_all_elements_located((By.CSS_SELECTOR, 'span[id^="jv-job-categories-codes-result-"]'))
                )
                vacancy = ' - '.join([span.text.strip() for span in vacancy_spans])
            except:
                pass

            # Lokasyon
            try:
                location_elem = wait.until(EC.presence_of_element_located((By.ID, "jv-details-job-location")))
                job_location = location_elem.text.strip()
            except:
                pass

            # İşveren Adı
            try:
                employer_name_elem = wait.until(EC.presence_of_element_located((By.ID, "jv-details-employer-name")))
                employer_name = employer_name_elem.text.strip()
            except:
                pass

            # Sektör
            try:
                job_sector_elem = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, 'dd.ecl-description-list__definition span#jv-employer-sector-codes-result')))
                job_sector = job_sector_elem.text.strip()
            except:
                pass

            # Şirket Adı
            try:
                company_name_elem = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, 'h2.ecl-u-type-heading-2.ecl-u-mt-s')))
                company_name = company_name_elem.text.strip()
            except:
                pass

            # İrtibat Kişi Adı
            try:
                contact_name_elem = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, 'li#jv-details-displayName-0')))
                contact_name = contact_name_elem.text.strip()
            except:
                pass

            # Excel'e yaz
            ws.append([
                email_address,
                vacancy,
                employer_name,
                contact_name,
                phone_number,
                company_name,
                job_location,
                job_sector,
            ])

            # Artık işlenen linki DataFrame'den silelim (Opsiyonel)
            df.drop(index, inplace=True)

            basarili_islem += 1
            print(sari_metin(f"[ + ] Veri kazıldı! Toplam Başarılı İşlem: {basarili_islem}"))

        except Exception as e:
            print(kirmizi_metin(f"Hata: {url} => {e}"))

        wb.save(excel_file)

    # Kalan linkleri tekrar yaz (Opsiyonel)
    df.to_csv(links_csv, index=False, header=False)

    driver.quit()

    end_time = time.time()
    elapsed_time = end_time - start_time
    print(yesil_metin(f"\nBütün Veriler Excel Dosyasına Kaydedildi! Toplam Başarılı İşlem: {basarili_islem}, Süre: {elapsed_time:.2f} sn"))

    return basarili_islem